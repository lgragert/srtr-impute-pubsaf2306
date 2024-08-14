import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import hlagenie

genie = hlagenie.init("3510", imputed=True)


# HLAGenie to get specific AA for each haplotype
def aminoacid(haplo_allele, which_allele):
    if which_allele == "DRB1":
        AA_list = [16, 28, 30, 32, 37, 47, 57, 60, 67, 70, 71, 74, 85, 86]
    if which_allele == "DQB1":
        AA_list = [9, 13, 23, 26, 37, 38, 57, 66, 67, 70, 86, 87]

    for AA in AA_list:
        D_B1_AA = []
        for row in range(len(haplo_allele)):
            allele = haplo_allele[which_allele][row]
            D_B1_AA.append(genie.getAA(allele, AA))
        haplo_allele[which_allele + "_" + str(AA)] = D_B1_AA

    return haplo_allele


# Transpose dataset to get the data to be what you want
def get_allele_format(optn_ag, ag_alleles, which_allele):

    # Only keep the first ten DR alleles for now
    ag_alleles = ag_alleles['IMGT/HLA alleles'].str.split(pat='/', expand=True)
    ag_alleles = pd.concat([optn_ag, ag_alleles], axis=1).reset_index(drop=True)

    # Want the rows to be indexed by allele
    transpose_ag_allele = ag_alleles.transpose()
    transpose_ag_allele.columns = transpose_ag_allele.iloc[0]
    transpose_ag_allele = transpose_ag_allele.iloc[1:]
    melt_ag_allele = pd.melt(transpose_ag_allele)
    melt_ag_allele = melt_ag_allele.dropna(ignore_index=True)
    if which_allele == "DRB1":
        melt_ag_allele = melt_ag_allele.rename(columns={'value': 'DRB1'})
    else:
        melt_ag_allele = melt_ag_allele.rename(columns={'value': 'DQB1'})
    return melt_ag_allele


# For adding antigen group later
ag_to_alleles = pd.read_csv('OPTN_antigens_to_alleles_CPRA.txt', sep='\t')
DR_optn_ag = ag_to_alleles[(ag_to_alleles['IMGT/HLA alleles'].str.contains('DRB1'))]
DQ_optn_ag = ag_to_alleles[(ag_to_alleles['IMGT/HLA alleles'].str.contains('DQB1'))]
DR_optn_antigen = DR_optn_ag['OPTN_Antigen']
DR_optn_ag_list = DR_optn_antigen.values.tolist()
DQ_optn_antigen = DQ_optn_ag['OPTN_Antigen']
DQ_optn_ag_list = DQ_optn_antigen.values.tolist()


DRB1_alleles = get_allele_format(DR_optn_antigen, ag_to_alleles, 'DRB1')
DQB1_alleles = get_allele_format(DQ_optn_antigen, ag_to_alleles, 'DQB1')


pops = ['AFA', 'API', 'CAU', 'HIS', 'MLT', 'NAM']
# pops = ['AFA']
haplotypes = pd.DataFrame()
for pop in pops:
    filename = 'freqs.' + pop + '.csv'
    freqs_file = pd.read_csv(filename)

    freqs_file = freqs_file.sort_values(by='Freq', ascending=False).reset_index(drop=True)
    freqs_file = freqs_file.iloc[:100, :]                                                                           # keep top 100 haplotypes
    freqs_file[['DRB345', 'DRB1', 'DQA1', 'DQB1']] = freqs_file['Haplo'].str.split(pat='~', expand=True)
    print(freqs_file)

    haplotypes = aminoacid(freqs_file, "DRB1")
    haplotypes = haplotypes.drop(['Haplo', 'Count', 'Freq', 'DQA1', 'DQB1'], axis=1)   # Formatting purposes

    haplotypes = pd.concat([haplotypes, freqs_file[['DQA1', 'DQB1']]], axis=1)

    # Create G1/G2 column for the DQB1
    g1g2_dict = {"DQB1*02": "G1", "DQB1*03": "G1", "DQB1*04": "G1", # grey
                 "DQB1*05": "G2", "DQB1*06": "G2"}   # pink

    g1g2_list = []
    for row in range(len(haplotypes)):
        dqb1 = haplotypes['DQB1'][row].split(sep=':')
        # dqa1 = haplotypes['DQA1'][row].split(sep=':')
        if dqb1[0] in g1g2_dict.keys():
            g1g2_list.append(g1g2_dict.get(dqb1[0]))
    haplotypes['G1/G2 Group'] = g1g2_list

    # Put aminoacid list after G1/G2 group classification
    haplotypes = aminoacid(haplotypes, 'DQB1')
    haplotypes = pd.concat([haplotypes, freqs_file[['Freq']]], axis=1)
    # Add DR Antigen Group
    drdq_col = haplotypes[['DRB345', 'DRB1', 'DQA1', 'DQB1']]
    DR_merge = pd.merge(drdq_col, DRB1_alleles, how='left', on='DRB1').drop_duplicates(keep='first', subset=['DRB345', 'DRB1', 'DQA1', 'DQB1'], ignore_index=True)
    print(DR_merge)
    DR_merge = DR_merge.rename(columns={'OPTN_Antigen': 'DR Antigen'})
    haplotypes = pd.concat([DR_merge['DR Antigen'], haplotypes], axis=1)
    # Add DQ Antigen Group
    DQ_merge = pd.merge(drdq_col, DQB1_alleles, how='left', on='DQB1').drop_duplicates(keep='last', subset=['DRB345', 'DRB1', 'DQA1', 'DQB1'], ignore_index=True)
    DQ_merge = DQ_merge.rename(columns={'OPTN_Antigen': 'DQ Antigen'})
    haplotypes = pd.concat([DQ_merge['DQ Antigen'], haplotypes], axis=1)

    rank_df = pd.DataFrame({"Rank": list(range(1, 101))})
    haplotypes = pd.concat([rank_df, haplotypes], axis=1)

    # Sort based on antigen/allele level DR edition
    # haplotypes = haplotypes.sort_values(by=['DRB345', 'DR Antigen', 'DRB1']).reset_index(drop=True)
    # Sort based on antigen/allele level DQ edition
    haplotypes = haplotypes.sort_values(by=['G1/G2 Group', 'DQ Antigen', 'DQB1', 'DQA1']).reset_index(drop=True)

    # Add blank lines based on how you're sorting
    for ser in DQ_optn_ag_list:
        if ser not in haplotypes['DQ Antigen'].values:
            continue
        else:
            index_list = haplotypes[(haplotypes["DQ Antigen"] == ser)].index.tolist()
            last_index = index_list[-1]
            blank_line = pd.DataFrame(index=[0, 1], columns=haplotypes.columns)
            haplotypes = pd.concat([haplotypes.loc[:last_index, :], blank_line, haplotypes.loc[last_index + 1:, :]], ignore_index=True)
    haplotypes = haplotypes[:-2]

    print(haplotypes)

    # Name changes depending on how you are sorting
    new_filename = "ClassII_AA_haplotype_patterns_DQ_" + pop + ".xlsx"
    haplotypes.to_excel(new_filename, index=False, sheet_name='Top ' + pop + ' Haplotypes')


def amino_acid_color(worksheet):
    max_column = worksheet.max_column
    for col in range(2, max_column + 1):
        column_letter = worksheet.cell(row=2, column=col).column_letter
        for cell in worksheet[column_letter]:
            if cell.value in AA_colors.keys():
                color_AA = PatternFill(patternType='solid', fgColor=AA_colors.get(cell.value))
                cell.fill = color_AA
            elif cell.value == "G1":
                color_g1 = PatternFill(patternType='solid', fgColor="ff8d8d8d")
                cell.fill = color_g1
            elif cell.value == "G2":
                color_g2 = PatternFill(patternType='solid', fgColor='00FF99CC')
                cell.fill = color_g2
            else:
                continue

    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['T'].width = 12
    worksheet.column_dimensions['U'].width = 12
    worksheet.column_dimensions['V'].width = 12

    return worksheet


# Spring color scheme
AA_colors = {"A": "ffababab", "C": "ff0893fe", "D": "fffc5977", "E": "ffff9c83", "F": "ff02f8b7", "G": "fff66c17",
             "H": "ffd2b907", "I": "ff47beff", "K": "ffffc56d", "L": "ff358b8f", "M": "ff4b9e93", "N": "ffeb684b",
             "P": "fffd75cf", "Q": "ffdeb16f", "R": "ffb38b2b", "S": "ffbb8187", "T": "ffd3add1", "V": "ff519db9",
             "W": "ff00d619", "Y": "ff38a549"}

for pop in pops:
    wb = load_workbook("ClassII_AA_haplotype_patterns_DQ_" + pop + ".xlsx")
    ws = wb['Top ' + pop + ' Haplotypes']

    ws = amino_acid_color(ws)
    wb.save(filename="ClassII_AA_haplotype_patterns_DQ_" + pop + ".xlsx")




