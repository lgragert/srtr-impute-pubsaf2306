import pandas as pd
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl import load_workbook

# Which DRB1 and DQB1 AA-MM are permissive and which are frequently mismatched with 1-DR Ag-MM and/or 1-DQ Ag-MM

# Permissive Ratio is AAMM within Low Risk : AAMM within High Risk Category
def permissiveness_ratio(low, high):
    if high == 0:
        permissive_ratio = str(low) + ":0"
    else:
        permissive_ratio = float(format(low / high, '.2f'))

    return permissive_ratio


# Ag-MM within Low and High Risk Categories either in overall or specific AAMM
def antigen_count(overall_low, overall_high, low_risk, high_risk, aa_or_pop, overall):

    if overall != True:
        low_ag_1_DR = len(overall_low[overall_low['REC_DR_MM_EQUIV_CUR'] == 1])
        low_ag_1_DQ = len(overall_low[overall_low['REC_DQ_MM_EQUIV_CUR'] == 1])
        low_both_1ag = len(overall_low[(overall_low['REC_DR_MM_EQUIV_CUR'] == 1) &
                                       (overall_low['REC_DQ_MM_EQUIV_CUR'] == 1)])
        low_ag_1_2_DR = len(overall_low[(overall_low['REC_DR_MM_EQUIV_CUR'] >= 1)])
        low_ag_1_2_DQ = len(overall_low[overall_low['REC_DQ_MM_EQUIV_CUR'] >= 1])
        low_ag_1_2_DRDQ = len(overall_low[(overall_low['REC_DR_MM_EQUIV_CUR'] >= 1) &
                                           (overall_low['REC_DQ_MM_EQUIV_CUR'] >= 1)])
        high_ag_1_DR = len(overall_high[overall_high['REC_DR_MM_EQUIV_CUR'] == 1])
        high_ag_1_DQ = len(overall_high[overall_high['REC_DQ_MM_EQUIV_CUR'] == 1])
        high_both_1ag = len(overall_high[(overall_high['REC_DR_MM_EQUIV_CUR'] == 1) &
                                         (overall_high['REC_DQ_MM_EQUIV_CUR'] == 1)])
        high_ag_1_2_DR = len(overall_high[(overall_high['REC_DR_MM_EQUIV_CUR'] >= 1)])
        high_ag_1_2_DQ = len(overall_high[overall_high['REC_DQ_MM_EQUIV_CUR'] >= 1])
        high_ag_1_2_DRDQ = len(overall_high[(overall_high['REC_DR_MM_EQUIV_CUR'] >= 1) &
                                            (overall_high['REC_DQ_MM_EQUIV_CUR'] >= 1)])
    else:
        low_ag_1_DR = low_ag_1_DQ = low_both_1ag = low_ag_1_2_DR = low_ag_1_2_DQ = low_ag_1_2_DRDQ = overall_low
        high_ag_1_DR = high_ag_1_DQ = high_both_1ag = high_ag_1_2_DR = high_ag_1_2_DQ = high_ag_1_2_DRDQ = overall_high


    # Ag-MM within Low Risk Category
    low_ag_DR1MM_count = len(low_risk[low_risk['REC_DR_MM_EQUIV_CUR'] == 1])
    low_ag_DR1MM_freq = float(format(low_ag_DR1MM_count / low_ag_1_DR, '.4f'))
    low_ag_DQ1MM_count = len(low_risk[low_risk['REC_DQ_MM_EQUIV_CUR'] == 1])
    low_ag_DQ1MM_freq = float(format(low_ag_DQ1MM_count / low_ag_1_DQ, '.4f'))
    low_ag_DR_DQ1MM_count = len(low_risk[(low_risk['REC_DR_MM_EQUIV_CUR'] == 1) &
                                         (low_risk['REC_DQ_MM_EQUIV_CUR'] == 1)])
    low_ag_DR_DQ1MM_freq = float(format(low_ag_DR_DQ1MM_count / low_both_1ag, '.4f'))
    low_ag_DR_1_2MM_count = len(low_risk[low_risk['REC_DR_MM_EQUIV_CUR'] >= 1])
    low_ag_DR_1_2MM_freq = float(format(low_ag_DR_1_2MM_count / low_ag_1_2_DR, '.4f'))
    low_ag_DQ_1_2MM_count = len(low_risk[low_risk['REC_DQ_MM_EQUIV_CUR'] >= 1])
    low_ag_DQ_1_2MM_freq = float(format(low_ag_DQ_1_2MM_count / low_ag_1_2_DQ, '.4f'))
    low_ag_DRDQ_1_2MM_count = len(low_risk[(low_risk['REC_DR_MM_EQUIV_CUR'] >= 1) &
                                           (low_risk['REC_DQ_MM_EQUIV_CUR'] >= 1)])
    low_ag_DRDQ_1_2MM_freq = float(format(low_ag_DRDQ_1_2MM_count / low_ag_1_2_DRDQ, '.4f'))

    # Ag-MM within High Risk Category
    high_ag_DR1MM_count = len(high_risk[high_risk['REC_DR_MM_EQUIV_CUR'] == 1])
    high_ag_DR1MM_freq = float(format(high_ag_DR1MM_count / high_ag_1_DR, '.4f'))
    high_ag_DQ1MM_count = len(high_risk[high_risk['REC_DQ_MM_EQUIV_CUR'] == 1])
    high_ag_DQ1MM_freq = float(format(high_ag_DQ1MM_count / high_ag_1_DQ, '.4f'))
    high_ag_DR_DQ1MM_count = len(high_risk[(high_risk['REC_DR_MM_EQUIV_CUR'] == 1) &
                                                     (high_risk['REC_DQ_MM_EQUIV_CUR'] == 1)])
    high_ag_DR_DQ1MM_freq = float(format(high_ag_DR_DQ1MM_count / high_both_1ag, '.4f'))
    high_ag_DR_1_2MM_count = len(high_risk[high_risk['REC_DR_MM_EQUIV_CUR'] >= 1])
    high_ag_DR_1_2MM_freq = float(format(high_ag_DR_1_2MM_count / high_ag_1_2_DR, '.4f'))
    high_ag_DQ_1_2MM_count = len(high_risk[high_risk['REC_DQ_MM_EQUIV_CUR'] >= 1])
    high_ag_DQ_1_2MM_freq = float(format(high_ag_DQ_1_2MM_count / high_ag_1_2_DQ, '.4f'))
    high_ag_DRDQ_1_2MM_count = len(high_risk[(high_risk['REC_DR_MM_EQUIV_CUR'] >= 1) &
                                             (high_risk['REC_DQ_MM_EQUIV_CUR'] >= 1)])
    high_ag_DRDQ_1_2MM_freq = float(format(high_ag_DRDQ_1_2MM_count / high_ag_1_2_DRDQ, '.4f'))

    # Relative Permissiveness is AAMM within Low Risk % - AAMM within High Risk %
    relative_permissive_DR = low_ag_DR1MM_freq - high_ag_DR1MM_freq
    relative_permissive_DQ = low_ag_DQ1MM_freq - high_ag_DQ1MM_freq
    relative_permissive_BOTH = low_ag_DR_DQ1MM_freq - high_ag_DR_DQ1MM_freq
    relative_permissive_1_2DR = low_ag_DR_1_2MM_freq - high_ag_DR_1_2MM_freq
    relative_permissive_1_2DQ = low_ag_DQ_1_2MM_freq - high_ag_DQ_1_2MM_freq
    relative_permissive_1_2DRDQ = low_ag_DRDQ_1_2MM_freq - high_ag_DRDQ_1_2MM_freq
    # Permissiveness Ratio is AAMM within Low Risk : AAMM within High Risk
    permissive_ratio_DR = permissiveness_ratio(low_ag_DR1MM_freq, high_ag_DR1MM_freq)
    permissive_ratio_DQ = permissiveness_ratio(low_ag_DQ1MM_freq, high_ag_DQ1MM_freq)
    permissive_ratio_BOTH = permissiveness_ratio(low_ag_DR_DQ1MM_freq, high_ag_DR_DQ1MM_freq)
    permissive_ratio_1_2DR = permissiveness_ratio(low_ag_DR_1_2MM_freq, high_ag_DR_1_2MM_freq)
    permissive_ratio_1_2DQ = permissiveness_ratio(low_ag_DQ_1_2MM_freq, high_ag_DQ_1_2MM_freq)
    permissive_ratio_1_2DRDQ = permissiveness_ratio(low_ag_DRDQ_1_2MM_freq, high_ag_DRDQ_1_2MM_freq)

    if overall != True:
        title = aa_or_pop + "_AAMM"
    else:
        title = "Total of " + aa_or_pop + " Population in SRTR Kidney"

    ag_risk_count = {title: [low_ag_DR1MM_count, low_ag_DR1MM_freq, high_ag_DR1MM_count, high_ag_DR1MM_freq,
                             permissive_ratio_DR, relative_permissive_DR, low_ag_DQ1MM_count, low_ag_DQ1MM_freq,
                             high_ag_DQ1MM_count, high_ag_DQ1MM_freq, permissive_ratio_DQ, relative_permissive_DQ,
                             low_ag_DR_DQ1MM_count, low_ag_DR_DQ1MM_freq, high_ag_DR_DQ1MM_count,
                             high_ag_DR_DQ1MM_freq, permissive_ratio_BOTH, relative_permissive_BOTH,
                             low_ag_DR_1_2MM_count, low_ag_DR_1_2MM_freq, high_ag_DR_1_2MM_count,
                             high_ag_DR_1_2MM_freq, permissive_ratio_1_2DR, relative_permissive_1_2DR,
                             low_ag_DQ_1_2MM_count, low_ag_DQ_1_2MM_freq, high_ag_DQ_1_2MM_count, high_ag_DQ_1_2MM_freq,
                             permissive_ratio_1_2DQ, relative_permissive_1_2DQ, low_ag_DRDQ_1_2MM_count, low_ag_DRDQ_1_2MM_freq,
                             high_ag_DRDQ_1_2MM_count, high_ag_DRDQ_1_2MM_freq, permissive_ratio_1_2DRDQ, relative_permissive_1_2DRDQ]}

    ag_df = pd.DataFrame.from_dict(ag_risk_count,
                                   columns=['Count LOW/1-DR Ag-MM', 'LOW/1-DR Ag-MM(%)', 'Count HIGH/1-DR Ag-MM',
                                            'HIGH/1-DR Ag-MM(%)', 'Permissiveness Ratio 1DR',
                                            'Relative Permissiveness 1DR', 'Count LOW/1-DQ Ag-MM', 'LOW/1-DQ Ag-MM(%)',
                                            'Count HIGH/1-DQ Ag-MM', 'HIGH/1-DQ Ag-MM(%)', 'Permissiveness Ratio 1DQ',
                                            'Relative Permissiveness 1DQ', 'Count LOW/1-DR/1-DQ Ag-MM',
                                            'LOW/1-DR/1-DQ Ag-MM(%)', 'Count HIGH/1-DR/1-DQ Ag-MM',
                                            'HIGH/1-DR/1-DQ Ag-MM(%)', 'Permissiveness Ratio 1-DR/1-DQ',
                                            'Relative Permissiveness 1-DR/1-DQ', 'Count LOW/>=1-DR Ag-MM',
                                            'LOW/>=1-DR Ag-MM(%)', 'Count HIGH/>=1-DR Ag-MM', 'HIGH/>=1-DR Ag-MM(%)',
                                            'Permissiveness Ratio >=1DR', 'Relative Permissiveness >=1DR',
                                            'Count LOW/>=1-DQ Ag-MM', 'LOW/>=1-DQ Ag-MM(%)', 'Count HIGH/>=1-DQ Ag-MM',
                                            'HIGH/>=1-DQ Ag-MM(%)', 'Permissiveness Ratio >=1DQ', 'Relative Permissiveness >=1DQ',
                                            'Count LOW/>=1-DR/>=1-DQ Ag-MM', 'LOW/>=1-DR/>=1-DQ Ag-MM(%)', 'Count HIGH/>=1-DR/>=1-DQ Ag-MM',
                                            'HIGH/>=1-DR/>=1-DQ Ag-MM(%)', 'Permissiveness Ratio >=1-DR/DQ', 'Relative Permissiveness >=1-DR/DQ'
                                            ],
                                   orient='index')

    return ag_df
    

def specific_amino_acid(SRTR, low_risk, high_risk, amino_acid):
    # Initialize SRTR to have specific AA-MM present
    col_AA = 'MM_' + amino_acid
    SRTR_AA = SRTR[(SRTR[col_AA] >= 1)]

    # Count specific AAMM in both SRTR and the Low Risk Category
    AAMM_low = low_risk[low_risk[col_AA].isin(SRTR_AA[col_AA])]
    low_count = len(AAMM_low)
    AAMM_low_count = float(format(low_count / len(low_risk), ".4f"))

    # Count Specific AAMM in both SRTR and the High Risk Category
    AAMM_high = high_risk[high_risk[col_AA].isin(SRTR_AA[col_AA])]
    high_count = len(AAMM_high)
    AAMM_high_count = float(format(high_count / len(high_risk), '.4f'))

    # Relative Permissiveness is AAMM within Low Risk % - AAMM within High Risk %
    relative_permissive = AAMM_low_count - AAMM_high_count
    # Permissiveness Ratio is AAMM within Low Risk : AAMM within High Risk
    permissive_ratio = permissiveness_ratio(low_count, high_count)

    AA_dict = {amino_acid + "_AAMM": [low_count, AAMM_low_count, high_count, AAMM_high_count, permissive_ratio, relative_permissive]}
    AA_df = pd.DataFrame.from_dict(AA_dict, columns=["Count LOW FIBERS", "LOW FIBERS(%)", "Count HIGH FIBERS", "HIGH FIBERS(%)", 'Permissiveness Ratio', 'Relative Permissiveness'], orient='index')

    # Count specific AAMM with certain Ag-MMs
    ag_risk_count_df = antigen_count(low_risk, high_risk, AAMM_low, AAMM_high, amino_acid, False)

    AA_df = pd.concat([AA_df, ag_risk_count_df], axis=1)

    return AA_df


# Fix to replicate 1
replicate = 1
SRTR_imputation_replicate_filename = "C:/Users/anpay/OneDrive/Documents/SRTR_AA_MM_matrix_grffail_" + str(replicate) + ".txt"
SRTR_df = pd.read_csv(SRTR_imputation_replicate_filename, sep='\t')

columns = ["PX_ID", "CAN_RACE", "REC_A_MM_EQUIV_CUR", "REC_B_MM_EQUIV_CUR", "REC_DR_MM_EQUIV_CUR", "REC_DQ_MM_EQUIV_CUR"]
DRB1 = []
DQB1 = []
# Create column names of SRTR and AA positions we will call for later
for num in range(6, 94):
    DRB1 += ['DRB1_' + str(num)]
    columns.append('MM_DRB1_' + str(num))
for num in range(6, 92):
    DQB1 += ['DQB1_' + str(num)]
    columns.append('MM_DQB1_' + str(num))

SRTR_df = SRTR_df[columns]
# pd.set_option('display.max_columns', None)
# print(SRTR_df.head())

# Fills all NA with 99 as 99 is not in the dataset
SRTR_df['REC_DR_MM_EQUIV_CUR'] = SRTR_df['REC_DR_MM_EQUIV_CUR'].fillna(99).astype(int)
SRTR_df['REC_DQ_MM_EQUIV_CUR'] = SRTR_df['REC_DQ_MM_EQUIV_CUR'].fillna(99).astype(int)
SRTR_df = SRTR_df.astype({'REC_DR_MM_EQUIV_CUR': 'int', 'REC_DQ_MM_EQUIV_CUR': 'int'})

# Changes the Ranking and Excel File Name Based on a Different (%) Column
changing_col = 'LOW FIBERS(%)'

if changing_col == 'LOW FIBERS(%)':
    excel_name = 'DR_DQ_AAMM_permissive_'
    A5_value = 'Count and Percentage of AA-MM within FIBERS Risk Cases, Sorted by AA-MM Frequency (Includes positions with >=0.5% AA-MM within Low FIBERS Freq)'


pops = ["ALLPOPS", "AFA", "CAU", "HIS", "HPI", "NAM", "ASI", "Multi-Racial"]
for pop in pops:

    DRB1_df = pd.DataFrame()
    overall_DRB1 = pd.DataFrame()
    DQB1_df = pd.DataFrame()
    overall_DQB1 = pd.DataFrame()

    if pop == "ALLPOPS":
        SRTR = SRTR_df
    else:
        SRTR = SRTR_df[(SRTR_df['CAN_RACE'] == pop)]

    total_pop = len(SRTR)
    total_pop_df = pd.DataFrame.from_dict({"Total of " + pop + " Population in SRTR Kidney": [total_pop, total_pop / total_pop]}, columns=['Count OVERALL', 'OVERALL(%)'], orient='index')

    # Fixate on Top Bin 9 for Consensus FIBERS High and Low Risk Categories
    low_risk_bin_9 = SRTR[(SRTR['MM_DQB1_30'] == 0) & (SRTR['MM_DQB1_55'] == 0) &
                          (SRTR['MM_DRB1_13'] == 0) & (SRTR['MM_DRB1_26'] == 0)]
    high_risk_bin_9 = SRTR[(SRTR['MM_DQB1_30'] >= 1) | (SRTR['MM_DQB1_55'] >= 1) |
                           (SRTR['MM_DRB1_13'] >= 1) | (SRTR['MM_DRB1_26'] >= 1)]
    count_low_risk = len(low_risk_bin_9)
    freq_low_risk = float(format(count_low_risk / total_pop, '.4f'))
    count_high_risk = len(high_risk_bin_9)
    freq_high_risk = float(format(count_high_risk / total_pop, '.4f'))

    # Relative Permissiveness and Permissiveness Ratio
    relative_permissive = freq_low_risk - freq_high_risk
    ratio_permissive = permissiveness_ratio(count_low_risk, count_high_risk)

    # Total amount in Dataset
    print("Total of " + pop + " in FIBERS Low Risk: ", count_low_risk)
    total_risk = {"Total of " + pop + " Population in SRTR Kidney": [count_low_risk, freq_low_risk, count_high_risk, freq_high_risk, ratio_permissive, relative_permissive]}
    total_risk_df = pd.DataFrame.from_dict(total_risk, columns=["Count LOW FIBERS", "LOW FIBERS(%)", "Count HIGH FIBERS", "HIGH FIBERS(%)", 'Permissiveness Ratio', 'Relative Permissiveness'], orient='index')
    both_totals = pd.concat([total_pop_df, total_risk_df], axis=1)

    # Overall totals of Ag-MM within Low Risk Groups
    ag_all_risk_count_df = antigen_count(count_low_risk, count_high_risk, low_risk_bin_9, high_risk_bin_9, pop, True)
    all_totals = pd.concat([both_totals, ag_all_risk_count_df], axis=1)

    # Look at DRB1 and DQB1 AA positions and find mismatches within the FIBERS Low Risk Category
    for AA in DRB1:
        DRB1_AAMM_df = specific_amino_acid(SRTR, low_risk_bin_9, high_risk_bin_9, AA)
        DRB1_df = pd.concat([DRB1_AAMM_df, DRB1_df])

        # Count how many DRB1 AA-MMs exist in dataset
        overall_drb1_dict = {AA + "_AAMM": [len(SRTR[(SRTR['MM_' + AA] >= 1)]), float(format(len(SRTR[(SRTR['MM_' + AA] >= 1)]) / total_pop, '.4f'))]}
        overall_drb1_df = pd.DataFrame.from_dict(overall_drb1_dict, columns=['Count OVERALL', 'OVERALL(%)'], orient='index')
        overall_DRB1 = pd.concat([overall_drb1_df, overall_DRB1])

    DRB1_df = pd.concat([DRB1_df, overall_DRB1], axis=1)

    for AA in DQB1:
        DQB1_AAMM_df = specific_amino_acid(SRTR, low_risk_bin_9, high_risk_bin_9, AA)
        DQB1_df = pd.concat([DQB1_AAMM_df, DQB1_df])

        # Count how many DRB1 AA-MMs exist in dataset
        overall_dqb1_dict = {AA + "_AAMM": [len(SRTR[(SRTR['MM_' + AA] >= 1)]), float(format(len(SRTR[(SRTR['MM_' + AA] >= 1)]) / total_pop, '.4f'))]}
        overall_dqb1_df = pd.DataFrame.from_dict(overall_dqb1_dict, columns=['Count OVERALL', 'OVERALL(%)'], orient='index')
        overall_DQB1 = pd.concat([overall_dqb1_df, overall_DQB1])
                                                                             
    DQB1_df = pd.concat([DQB1_df, overall_DQB1], axis=1)

    DRB1_DQB1_AA = pd.concat([DRB1_df, DQB1_df])

    # Drop rows that are < 0.05 % in the frequency we are observing
    DRB1_df = DRB1_df[(DRB1_df[changing_col] >= 0.005)]
    DQB1_df = DQB1_df[(DQB1_df[changing_col] >= 0.005)]
    DRB1_DQB1_AA = DRB1_DQB1_AA[(DRB1_DQB1_AA[changing_col] >= 0.005)]

    # Sort Values from highest to lowest for each subset
    DRB1_df = DRB1_df.sort_values(by=changing_col, ascending=False)
    DQB1_df = DQB1_df.sort_values(by=changing_col, ascending=False)
    DRB1_DQB1_AA = DRB1_DQB1_AA.sort_values(by=changing_col, ascending=False)

    # Append the total count to the top so that we know how to calculate the frequencies
    DRB1_df = pd.concat([all_totals, DRB1_df])
    DQB1_df = pd.concat([all_totals, DQB1_df])
    DRB1_DQB1_AA = pd.concat([all_totals, DRB1_DQB1_AA])

    
    with pd.ExcelWriter(excel_name + pop + "_tables.xlsx") as writer:
        DRB1_df.to_excel(writer, sheet_name='DRB1')
        DQB1_df.to_excel(writer, sheet_name='DQB1')
        DRB1_DQB1_AA.to_excel(writer, sheet_name='DRB1_DQB1')


# TRS added to the excel sheet
def trs_column(average_trs, excel_sheet):
    loci_only = excel_sheet
    loci_only.rename(columns={'Unnamed: 0':'Loci_AA'}, inplace=True)
    loci_only = excel_sheet['Loci_AA']
    
    trs_merge = pd.merge(average_trs, loci_only, how='right', on='Loci_AA')

    trs_merge = trs_merge['TRS_Average']
    excel_sheet.insert(9, 'Tying Resolution Score Average', trs_merge)

    return excel_sheet

# ALD added to the excel sheet
def ALD_FIBERS(ALD, excel_sheet, sfvt_annot):

    loci_only = excel_sheet
    loci_only.rename(columns={'Loci_AA': 'Loc-Pos'}, inplace=True)

    loci_only['Loc-Pos'] = loci_only['Loc-Pos'].str.replace('_', '-')
    loci_only['Loc-Pos'] = loci_only['Loc-Pos'].str.replace('-AAMM', '')
    loci_only = loci_only['Loc-Pos']

    DRB1_13 = ALD[['Loc-Pos', 'DRB1-13']]
    DRB1_26 = ALD[['Loc-Pos', 'DRB1-26']]
    DQB1_30 = ALD[['Loc-Pos', 'DQB1-30']]
    DQB1_55 = ALD[['Loc-Pos', 'DQB1-55']]
    
    ALD13_merge = pd.merge(DRB1_13, loci_only, how='right', on='Loc-Pos')
    ALD26_merge = pd.merge(DRB1_26, loci_only, how='right', on='Loc-Pos')
    ALD30_merge = pd.merge(DQB1_30, loci_only, how='right', on='Loc-Pos')
    ALD55_merge = pd.merge(DQB1_55, loci_only, how='right', on='Loc-Pos')

    ALD13_merge = ALD13_merge['DRB1-13']
    ALD26_merge = ALD26_merge['DRB1-26']
    ALD30_merge = ALD30_merge['DQB1-30']
    ALD55_merge = ALD55_merge['DQB1-55']

    excel_sheet.insert(10, 'ALD DRB1_13', ALD13_merge)
    excel_sheet.insert(11, 'ALD DRB1_26', ALD26_merge)
    excel_sheet.insert(12, 'ALD DQB1_30', ALD30_merge)
    excel_sheet.insert(13, 'ALD DQB1_55', ALD55_merge)

    sfvt_annot.rename(columns={'Loci_AA': 'Loc-Pos'}, inplace=True)

    sfvt_merge = pd.merge(sfvt_annot, loci_only, how='right', on='Loc-Pos')
    sfvt_merge = sfvt_merge.drop(columns=['Loc-Pos'])

    excel_sheet = pd.concat([excel_sheet, sfvt_merge], axis=1)

    excel_sheet.set_index('Loc-Pos', inplace=True)

    return excel_sheet


AA_trs = pd.read_csv("SRTR_HLA_AA_TRS_Average.csv")
sfvt_annot_pos = pd.read_csv('SFVT_annotation_string.csv', index_col=0)
# Concat SFVT Annotation to the end of every table and add config file for a "cheatsheet"
sf_config = pd.read_csv('SF_annot_mapping.cfg')
# TRS and ALD Additional Columns 
for pop in pops:

    if pop == "ALLPOPS":
        AA_TRS = AA_trs[(AA_trs['Subject_Type'] == 'SUBJECT') & (AA_trs["Ethnicity"] == 'ALL') & (AA_trs['Locus'].str.startswith("D"))]
    elif pop == 'Multi-Racial':
        AA_TRS = AA_trs[(AA_trs['Subject_Type'] == 'SUBJECT') & (AA_trs["Ethnicity"] == 'MLT') & (AA_trs['Locus'].str.startswith("D"))]
    else:
        AA_TRS = AA_trs[(AA_trs['Subject_Type'] == 'SUBJECT') & (AA_trs["Ethnicity"] == pop) & (AA_trs['Locus'].str.startswith("D"))]

    AA_TRS['Loci_AA'] = AA_TRS['Locus'] + "_" + AA_TRS['AA_Position'].astype(str) + "_AAMM"

    AA_TRS = AA_TRS[["Loci_AA", "TRS_Average"]]

    DRB1 = pd.read_excel(excel_name + pop + "_tables.xlsx", sheet_name='DRB1')
    DQB1 = pd.read_excel(excel_name + pop + "_tables.xlsx", sheet_name='DQB1')
    DRB1_DQB1 = pd.read_excel(excel_name + pop + "_tables.xlsx", sheet_name='DRB1_DQB1')

    DRB1_TRS = trs_column(AA_TRS, DRB1)
    DQB1_TRS = trs_column(AA_TRS, DQB1)
    DRB1_DQB1_TRS = trs_column(AA_TRS, DRB1_DQB1)

    if pop == 'AFA' or pop == 'CAU' or pop == 'API':
        ALD_df = pd.read_csv(pop + '.ALD.matrix.DQB1_DRB1.csv')
    else:
        ALD_df = pd.read_csv('HIS.ALD.matrix.DQB1_DRB1.csv')

    # We want DRB1 13, 26 and DQB1 30, 55

    DRB1_ALD = ALD_FIBERS(ALD_df, DRB1_TRS, sfvt_annot_pos)
    DQB1_ALD = ALD_FIBERS(ALD_df, DQB1_TRS, sfvt_annot_pos)
    DRB1_DQB1_ALD = ALD_FIBERS(ALD_df, DRB1_DQB1_TRS, sfvt_annot_pos)

    with pd.ExcelWriter(excel_name + pop + "_tables.xlsx") as writer:
        DRB1_ALD.to_excel(writer, sheet_name='DRB1')
        DQB1_ALD.to_excel(writer, sheet_name='DQB1')
        DRB1_DQB1_ALD.to_excel(writer, sheet_name='DRB1_DQB1')
        sf_config.to_excel(writer, sheet_name='SF_ANNOT INFO')



# Excel Spreadsheet Design and Adjustments
FIBERS_heading = NamedStyle(name="FIBERS_headings")
FIBERS_heading.font = Font(bold=True, color="00FF0000")


def excel_adjuster(worksheet, A5_values):

    max_column = worksheet.max_column

    for column_index in range(1, max_column + 1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = 24

        for cell in worksheet[column_letter]:
            cell.alignment = Alignment(horizontal='left')

    number_format = '0.00%'
    worksheet_col = ['C', 'E', 'G', 'I', 'P', 'R', 'T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH', 'AJ', 'AL', 'AN',  'AP', 'AR', 'AT', 'AV', 'AX']

    for col in worksheet_col:
        for cell in worksheet[col]:
            cell.number_format = number_format

    worksheet.column_dimensions['A'].width = 34
    worksheet.insert_rows(2)
    worksheet.insert_rows(4)

    worksheet['A1'].value = ''
    worksheet['A2'].value = "Totals Used for Frequencies(%)"
    worksheet['A4'].value = A5_values

    worksheet['A2'].style = FIBERS_heading
    worksheet['A4'].style = FIBERS_heading

    return worksheet


AA_trs = pd.read_csv("SRTR_HLA_AA_TRS_Average.csv")
for pop in pops:
    wb = load_workbook(filename=excel_name + pop + "_tables.xlsx")
    DR_ws = wb['DRB1']
    DQ_ws = wb['DQB1']
    both_ws = wb['DRB1_DQB1']
    sf_annot_ws = wb['SF_ANNOT INFO']

    excel_adjuster(DR_ws, A5_value)
    excel_adjuster(DQ_ws, A5_value)
    excel_adjuster(both_ws, A5_value)
    sf_annot_ws.delete_cols(1)
    sf_annot_ws.column_dimensions['A'].width = 83
    sf_annot_ws.column_dimensions['B'].width = 20

    wb.save(filename=excel_name + pop + "_tables.xlsx")



