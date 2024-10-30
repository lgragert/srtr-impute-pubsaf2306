import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook
from collections import defaultdict
import json
import matplotlib.pyplot as plt

# Script to generate summary stats for amino acid MM in replicate 1


# Get FIBERS High and Low risk groups for each bin (FIBERS1.0 threshold of 1 AAMM)
def high_low_bins(SRTR, aamm_config, bin_num):
    # Initialize low risk condition
    low_risk = (SRTR[("MM_" + aamm_config['FIBERS_BINS'][bin_num][0])] == 0)
    for aamm in aamm_config['FIBERS_BINS'][bin_num][1:]:
        low_risk &= (SRTR[("MM_" + aamm)] == 0)
    SRTR_Low_Risk = SRTR[low_risk]

    # Initialize high risk condition
    high_risk = (SRTR[("MM_" + aamm_config['FIBERS_BINS'][bin_num][0])] >= 1)
    for aamm in aamm_config['FIBERS_BINS'][bin_num][1:]:
        high_risk |= (SRTR[("MM_" + aamm)] >= 1)

    SRTR_High_Risk = SRTR[high_risk]
    return SRTR_Low_Risk, SRTR_High_Risk


# Get FIBERS High and Low risk groups for each bin (FIBERS2.0 bin-specific thresholds)
def high_low_bins_thresh(SRTR, aamm_config, bin_num, bin_param):
    bin_key = bin_num
    bin_positions = aamm_config['FIBERS_BINS'][bin_key]
    
    # Find the threshold for this bin
    threshold = next(item['Threshold'] for item in bin_param if item['FIBERS Bin'] == bin_key)

    # Calculate the sum of AAMM for the positions in this bin
    aamm_sum = SRTR[[f"MM_{aamm}" for aamm in bin_positions]].sum(axis=1)

    # Assign low risk if the sum is less than the threshold
    low_risk = (aamm_sum < threshold)
    SRTR_Low_Risk = SRTR[low_risk]

    # Assign high risk if the sum is greater than or equal to the threshold
    high_risk = (aamm_sum >= threshold)
    SRTR_High_Risk = SRTR[high_risk]

    return SRTR_Low_Risk, SRTR_High_Risk


# Format counts to have a 'count (percentage %)' for each excel cell
def format_counts(ini_count, total_sum, overall, dict_key=None):
    if overall is True:
        ag_perc = str(ini_count) + " (" + str(format(100 * ini_count / total_sum, '.2f')) + "%)"
    else:
        total_sum[0].replace(to_replace="0 ", value="1", inplace=True) # We cannot divide by 0, so this little trick helps with it
        ag_perc = str(ini_count) + " (" + str(format(100 * ini_count / int(total_sum.loc[dict_key][0]), '.2f')) + "%)"
    return ag_perc


# Overall Antigen Mismatch Counts Function
def antigen_count(low_high_bin, total_count, high_or_low, overall):
    # Initialize every row
    antigen_DR_NA_count = len(low_high_bin[low_high_bin['REC_DR_MM_EQUIV_CUR'] == 99])
    antigen_DQ_NA_count = len(low_high_bin[low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 99])

    antigen_ABDR_0MM_count = len(low_high_bin[(low_high_bin['REC_A_MM_EQUIV_CUR'] == 0) &
                                              (low_high_bin['REC_B_MM_EQUIV_CUR'] == 0) &
                                              (low_high_bin['REC_DR_MM_EQUIV_CUR'] == 0)])

    antigen_DR_0MM_count = len(low_high_bin[low_high_bin['REC_DR_MM_EQUIV_CUR'] == 0])
    antigen_DQ_0MM_count = len(low_high_bin[low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 0])

    antigen_DR_1MM_count = len(low_high_bin[low_high_bin['REC_DR_MM_EQUIV_CUR'] == 1])
    antigen_DQ_1MM_count = len(low_high_bin[low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 1])

    antigen_DR_2MM_count = len(low_high_bin[low_high_bin['REC_DR_MM_EQUIV_CUR'] == 2])
    antigen_DQ_2MM_count = len(low_high_bin[low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 2])

    antigen_DR_0MM_DQ_0MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 0) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 0)])

    antigen_DR_0MM_DQ_1MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 0) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 1)])

    antigen_DR_1MM_DQ_0MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 1) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 0)])

    antigen_DR_1MM_DQ_1MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 1) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 1)])

    antigen_DR_0MM_DQ_2MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 0) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 2)])

    antigen_DR_2MM_DQ_0MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 2) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 0)])

    antigen_DR_1MM_DQ_2MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 1) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 2)])

    antigen_DR_2MM_DQ_1MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 2) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 1)])

    antigen_DR_2MM_DQ_2MM_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 2) &
                                                   (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 2)])

    antigen_DR_0MM_DQ_NA_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 0) &
                                                  (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 99)])

    antigen_DR_1MM_DQ_NA_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 1) &
                                                  (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 99)])

    antigen_DR_2MM_DQ_NA_count = len(low_high_bin[(low_high_bin['REC_DR_MM_EQUIV_CUR'] == 2) &
                                                  (low_high_bin['REC_DQ_MM_EQUIV_CUR'] == 99)])
    # Make the rows depending on if overall count or not
    ag_count_dict = {"Missing-DR AgMM Count": antigen_DR_NA_count, "Missing-DQ AgMM Count": antigen_DQ_NA_count,
                      "0-ABDR AgMM Count": antigen_ABDR_0MM_count, "0-DR AgMM Count": antigen_DR_0MM_count,
                      "0-DQ AgMM Count": antigen_DQ_0MM_count, "1-DR AgMM Count": antigen_DR_1MM_count,
                      "1-DQ AgMM Count": antigen_DQ_1MM_count, "2-DR AgMM Count": antigen_DR_2MM_count,
                      "2-DQ AgMM Count": antigen_DQ_2MM_count, "0-DR and 0-DQ AgMM Count": antigen_DR_0MM_DQ_0MM_count,
                      "0-DR and 1-DQ AgMM Count": antigen_DR_0MM_DQ_1MM_count,"1-DR and 0-DQ AgMM Count":
                          antigen_DR_1MM_DQ_0MM_count, "1-DR and 1-DQ AgMM Count": antigen_DR_1MM_DQ_1MM_count,
                      "0-DR and 2-DQ AgMM Count": antigen_DR_0MM_DQ_2MM_count, "2-DR and 0-DQ AgMM Count":
                          antigen_DR_2MM_DQ_0MM_count, "1-DR and 2-DQ AgMM Count": antigen_DR_1MM_DQ_2MM_count,
                      "2-DR and 1-DQ AgMM Count": antigen_DR_2MM_DQ_1MM_count,"2-DR and 2-DQ AgMM Count":
                          antigen_DR_2MM_DQ_2MM_count, "0-DR and Missing-DQ AgMM Count": antigen_DR_0MM_DQ_NA_count,
                      "1-DR and Missing-DQ AgMM Count": antigen_DR_1MM_DQ_NA_count,
                      "2-DR and Missing-DQ AgMM Count": antigen_DR_2MM_DQ_NA_count}

    ag_perc_dict = {}
    for key, value in ag_count_dict.items():
        ag_perc = format_counts(value, total_count, overall, key)
        ag_perc_dict[key] = ag_perc

    if high_or_low == "high":
        # "Leave blank"'s are due to formatting issues, but they will be fixed later in the process
        # Join the leave blanks with the percentage dictionaries made previously
        high_header_rows = {"Leave blank": " ", "leave blank": " ", "AA-MM within Ag-MM and FIBERS High Risk Combos": ""}
        antigen_dict = {**high_header_rows, **ag_perc_dict}

    if high_or_low == "low":
        low_header_row = {"AA-MM within Ag-MM and FIBERS Low Risk Combos": ""}
        antigen_dict = {**low_header_row, **ag_perc_dict}

    return antigen_dict


# Specific Amino Acids Count Function: DRB1_13 _26 and DQB1_30 _55
def specific_amino_acid(SRTR, high_list, low_list, high_overall, low_overall, amino_acid):

    # Initialize SRTR to have specific AA-MM present
    col_AA = 'MM_' + amino_acid
    SRTR_AA = SRTR[(SRTR[col_AA] >= 1)]

    high_aa_ag_dict = defaultdict(lambda: defaultdict(dict))
    low_aa_ag_dict = defaultdict(lambda: defaultdict(dict))
    total_low_count_dict = defaultdict()
    total_high_count_dict = defaultdict()

    for idx, high in enumerate(high_list):
        # Checks to compare AA-MM presence in bin with high risk
        high_bin_SRTR = high[high[col_AA].isin(SRTR_AA[col_AA])]
        total_count = len(high)
        high_count = len(high_bin_SRTR)
        binSRTR_count = str(high_count) + " (" + str(format(100 * high_count / total_count, ".2f")) + "%)"
        total_high_count_dict["FIBERS_Bin_" + str(idx + 1)] = high_count
        # Need these totals for the denominators for the Ag-MM Counts
        no_percents_high = high_overall["FIBERS_Bin_" + str(idx + 1)].str.split("(", expand=True)
        no_percents_high = no_percents_high.drop(columns=1, index=["Leave blank", "leave blank",
                                                                   "AA-MM within Ag-MM and FIBERS High Risk Combos",
                                                                   "FIBERS High Risk Bin Thresh"])
        antigen_dict = antigen_count(high_bin_SRTR, no_percents_high, "high", False)
        high_aa_ag_dict["FIBERS_Bin_" + str(idx + 1)] = antigen_dict
        high_aa_ag_dict["FIBERS_Bin_" + str(idx + 1)]['FIBERS High Risk Bin Thresh'] = binSRTR_count

    for idx, low in enumerate(low_list):
        # Checks to compare AA-MM presence in bin with high risk
        low_bin_SRTR = low[low[col_AA].isin(SRTR_AA[col_AA])]
        total_count = len(low)
        low_count = len(low_bin_SRTR)
        binSRTR_count = str(low_count) + " (" + str(format(100 * low_count / total_count, ".2f")) + "%)"
        total_low_count_dict["FIBERS_Bin_" + str(idx + 1)] = low_count
        # Need these totals for the denominators for the Ag-MM Counts
        no_percents_low = low_overall["FIBERS_Bin_" + str(idx + 1)].str.split("(", expand=True)
        no_percents_low = no_percents_low.drop(columns=1, index=["AA-MM within Ag-MM and FIBERS Low Risk Combos",
                                                                 "FIBERS Low Risk Bin Thresh"])
        antigen_dict = antigen_count(low_bin_SRTR, no_percents_low, "low", False)
        low_aa_ag_dict["FIBERS_Bin_" + str(idx + 1)] = antigen_dict
        low_aa_ag_dict["FIBERS_Bin_" + str(idx + 1)]['FIBERS Low Risk Bin Thresh'] = binSRTR_count

    aa_high = pd.DataFrame.from_dict(high_aa_ag_dict, orient='columns')
    aa_low = pd.DataFrame.from_dict(low_aa_ag_dict, orient='columns')
    # Add a total AA-MM Count at the top of the spreadsheet
    total_low_count_df = pd.DataFrame.from_dict(total_low_count_dict, orient='index', columns=["Low Risk"])
    total_high_count_df = pd.DataFrame.from_dict(total_high_count_dict, orient='index', columns=["High Risk"])
    total_risk_count = pd.concat([total_low_count_df, total_high_count_df], axis=1)
    total_risk_count_df = pd.DataFrame()
    total_risk_count_df['Total_MM'] = total_risk_count.sum(axis=1)
    total_risk_count_df = total_risk_count_df.transpose()

    aa_ag_df = pd.concat([aa_high, aa_low])
    aa_ag_df = pd.concat([total_risk_count_df, aa_ag_df])

    aa_ag_df.name = amino_acid  # Will Create a Nested Dictionary with Named DataFrames

    return aa_ag_df


# AA count function to provide distribution of AAMM frequency across positions
def AA_total_count(SRTR, high_list, low_list, amino_acid):

    # Initialize SRTR to have specific AA-MM present
    col_AA = 'MM_' + amino_acid
    SRTR_AA = SRTR[(SRTR[col_AA] >= 1)]

    high_count = 0
    low_count = 0

    for idx, high in enumerate(high_list):
        # Checks to compare AA-MM presence in bin with high risk
        high_bin_SRTR = high[high[col_AA].isin(SRTR_AA[col_AA])]
        total_count = len(high)
        high_count = len(high_bin_SRTR)

    for idx, low in enumerate(low_list):
        # Checks to compare AA-MM presence in bin with high risk
        low_bin_SRTR = low[low[col_AA].isin(SRTR_AA[col_AA])]
        total_count = len(low)
        low_count = len(low_bin_SRTR)

    # total risk count for each AA
    #total_count = next(iter(total_high_count_dict.values())) + next(iter(total_low_count_dict.values()))
    total_count = high_count + low_count

    return total_count

###########################

# Handle FIBERS output file (from Harsh) to config file for script

# Read the FIBERS output file of new runs
FIBERS_output_file = pd.read_csv('Fibers2.0_hla_7locus_baseline_summary.csv')

# Read in the config file to update with the new bins from FIBERS
with open('amino_acids.json', 'r') as f:
    config = json.load(f)

# Create a list to store FIBERS run metadata for each bin
bin_param = []

# Process each row of the FIBERS output file
for index, row in FIBERS_output_file.iterrows():
    bin_number = index + 1  # Assuming rows are in order from 1 to 10 - FOR NEWIMP_1 REPLICATE 1 ONLY
    bin_key = f"Bin_{bin_number}"
    
    # Get the list of features from the output file
    features = eval(row['Bin Features'])
    cleaned_features = [feature.replace('MM_', '') for feature in features]
    
    # Update the bin dictionary in the config file
    if bin_key in config['FIBERS_BINS']:
        config['FIBERS_BINS'][bin_key] = cleaned_features

    # Store additional information for this bin
    bin_param.append({
        'FIBERS Bin': bin_key,
        'Dataset Filename': row['Dataset Filename'], 
        'Random Seed': row['Random Seed'],
        'Bin Features': row['Bin Features'],
        'Threshold': row['Threshold'],
        'Fitness': row['Fitness'], 
        'Pre-Fitness': row['Pre-Fitness'],
        'Log-Rank Score': row['Log-Rank Score'],
        'Log-Rank p-value': row['Log-Rank p-value'],
        'Bin Size': row['Bin Size'],
        'Group Ratio': row['Group Ratio'],
        'Count At/Below Threshold': row['Count At/Below Threshold'],
        'Count Above Threshold': row['Count Above Threshold'],
        'Birth Iteration': row['Birth Iteration'],
        'Deletion Probability': row['Deletion Probability'],
        'Cluster': row['Cluster'],
        'Residual': row['Residual'],
        'Residual p-value': row['Residual p-value'],
        'Unadjusted HR': row['Unadjusted HR'],
        'Unadjusted HR CI': row['Unadjusted HR CI'],
        'Unadjusted HR p-value': row['Unadjusted HR p-value'],
        'Adjusted HR': row['Adjusted HR'],
        'Adjusted HR CI': row['Adjusted HR CI'],
        'Adjusted HR p-value': row['Adjusted HR p-value'],
        'Runtime': row['Runtime']
        #'Features': ', '.join(cleaned_features)  # Join features into a single string
    })

# Save the updated config file
with open('amino_acids.json', 'w') as f:
    json.dump(config, f, indent=2)

print("CONFIG file has been updated.")

################################

# Create a dictionary of threshold values from `bin_param`
thresholds = {item['FIBERS Bin']: item['Threshold'] for item in bin_param}

################################

# Get AA in config file
with open('amino_acids.json', 'r') as file:
    config = json.load(file)

# Reference table of AA-MM in each bin
table_df = pd.DataFrame(config['FIBERS_BINS'].values(), index=config['FIBERS_BINS'])

# SRTR imputation replicate number
replicate = config.get('Imp_Replicate')
# Indicator for bin number for OVERALL and AA tabs
bin_name = config.get('Bin_of_Interest')

# Get AA of interest along with bin of AA that we found most interesting
# AAMM_tabs creates the tabs in the Excel Spreadsheet
# Last run, bin 9 was most interesting
aa_of_int = config['AA_of_Interest']
bin_of_int = config['FIBERS_BINS'][bin_name]
AAMM_tabs = aa_of_int + bin_of_int

# Remove duplicates to keep unique AAMMs
AAMM_tabs = list(set(AAMM_tabs))
AAMM_tabs.sort()

# SRTR using only one replicate/realization
SRTR_imputation_replicate_filename = "SRTR_AA_MM_9loc_matrix_DQAg_" + str(replicate) + ".txt.gz"
SRTR_df = pd.read_csv(SRTR_imputation_replicate_filename, sep='\t', compression='gzip')

# Subset amino acid MM and antigen MM columns from Dataframe
SRTR_ag = SRTR_df[["PX_ID", "CAN_RACE", "REC_A_MM_EQUIV_CUR", "REC_B_MM_EQUIV_CUR", "REC_DR_MM_EQUIV_CUR",
                   "REC_DQ_MM_EQUIV_CUR"]]
# The AAMM are all the ones found in the bins, so create a for loop to get list and get only unique AAMM
# Want to subset SRTR to only the AAMM found in all the bins and AA of interests
aa_list = []
for bin in config['FIBERS_BINS']:
    specific_bin = config['FIBERS_BINS'][bin]
    # Also want all the AA of interest too, so add to the list
    aa_of_int = config['AA_of_Interest']
    for aa in specific_bin:
        if aa not in aa_list:
            aa_list.append(aa)
    for aa in aa_of_int:
        if aa not in aa_list:
            aa_list.append(aa)

aa_list_header = ["MM_" + AAMM for AAMM in aa_list]
SRTR_aa = SRTR_df[aa_list_header]

# Append AA with the Ag/population group info DF
SRTR_df = pd.concat([SRTR_ag, SRTR_aa], axis=1)

# Convert AgMM to integers and handle NAs due to missing DQ typing
# 99 - Not tested per SRTR_df data dictionary
SRTR_df['REC_DR_MM_EQUIV_CUR'] = SRTR_df['REC_DR_MM_EQUIV_CUR'].fillna(99).astype(int)
SRTR_df['REC_DQ_MM_EQUIV_CUR'] = SRTR_df['REC_DQ_MM_EQUIV_CUR'].fillna(99).astype(int)
SRTR_df = SRTR_df.astype({'REC_DR_MM_EQUIV_CUR': 'int', 'REC_DQ_MM_EQUIV_CUR': 'int'})


# Loop through each population and overall
pops = ["OVERALL", 
        "          16: Black or African American", 
        "           8: White", 
        "        2000: Hispanic/Latino", 
        "         128: Native Hawaiian or Other Pacific Islander", 
        "          32: American Indian or Alaska Native", 
        "          64: Asian", 
        "OTHER"]

popnames = {
    'OVERALL': 'OVERALL', 
    'OTHER': 'OTHER', 
    '        2000: Hispanic/Latino': 'HIS',
    '          16: Black or African American': 'AFA',
    '           8: White': 'CAU',
    '          64: Asian': 'ASI',
    '          32: American Indian or Alaska Native': 'NAM',
    '         128: Native Hawaiian or Other Pacific Islander': 'HPI'
}

# Handle population groups with OTHER
SRTR_df['CAN_RACE'] = np.where(SRTR_df['CAN_RACE'].isin(pops), SRTR_df['CAN_RACE'], "OTHER")

# Create a new Excel workbook for datatables for AAMM frequency distribution across positions 
# Separate tab per population group
workbook = Workbook()
workbook.remove(workbook.active)  # Remove the default sheet

# Initialize all the bins and make nested dictionaries for each sheet
high_count_dict = defaultdict(lambda: defaultdict(dict))
low_count_dict = defaultdict(lambda: defaultdict(dict))
amino_acid_dict = defaultdict(lambda: defaultdict(dict))
AAMM_df = defaultdict(lambda: defaultdict(dict))

for pop in pops:
    print("Working on population: ", pop)

    popcode = popnames[pop]

    if pop == "OVERALL":
        SRTR = SRTR_df
    else:
        SRTR = SRTR_df[(SRTR_df['CAN_RACE'] == pop)]

    overall_SRTR_count = len(SRTR)
    print("Overall Count:", overall_SRTR_count)

    # Get high and low risk counts for each bin discovered by FIBERS
    low_bins = []
    high_bins = []
    for bin in config['FIBERS_BINS']:
        low_freq, high_freq = high_low_bins_thresh(SRTR, config, bin, bin_param)
        low_bins.append(low_freq)
        high_bins.append(high_freq)

    # For OVERALL Excel Sheet
    for idx, high in enumerate(high_bins):
        high_total = len(high)
        high_overall_count = str(high_total) + " (" + str(format(100 * high_total / overall_SRTR_count, ".2f")) + "%)"
        antigen_high = antigen_count(high, high_total, "high", True)
        high_count_dict["FIBERS_Bin_" + str(idx + 1)] = antigen_high
        high_count_dict["FIBERS_Bin_" + str(idx + 1)]["FIBERS High Risk Bin Thresh"] = high_overall_count

    for idx, low in enumerate(low_bins):
        low_total = len(low)
        low_overall_count = str(low_total) + " (" + str(format(100 * low_total / overall_SRTR_count, ".2f")) + "%)"
        antigen_low = antigen_count(low, low_total, "low", True)
        low_count_dict["FIBERS_Bin_" + str(idx + 1)] = antigen_low
        low_count_dict["FIBERS_Bin_" + str(idx + 1)]["FIBERS Low Risk Bin Thresh"] = low_overall_count

    overall_high = pd.DataFrame.from_dict(high_count_dict, orient='columns')
    overall_low = pd.DataFrame.from_dict(low_count_dict, orient='columns')
    OVERALL = pd.concat([overall_high, overall_low])
    # Adds a row for the total population in each bin
    pop_row = pd.DataFrame({"FIBERS_Bin_1": overall_SRTR_count, "FIBERS_Bin_2": overall_SRTR_count, "FIBERS_Bin_3": overall_SRTR_count, "FIBERS_Bin_4": overall_SRTR_count,
                            "FIBERS_Bin_5": overall_SRTR_count, "FIBERS_Bin_6": overall_SRTR_count, "FIBERS_Bin_7": overall_SRTR_count, "FIBERS_Bin_8": overall_SRTR_count,
                            "FIBERS_Bin_9": overall_SRTR_count, "FIBERS_Bin_10": overall_SRTR_count}, index=["Total " + popcode + " Population"])
    OVERALL = pd.concat([pop_row, OVERALL])

    # For Each Single AA-MM Excel Sheet
    for AA in AAMM_tabs:
        AAMM_df[AA] = specific_amino_acid(SRTR, high_bins, low_bins, overall_high, overall_low, AA)

    # pd.set_option('display.max_columns', None)
    # pd.set_option('display.max_colwidth', None)
    # print(AAMM_df)

    ####################

    # Initialize dictionary for AAs and their total counts
    AAMM_tot = {}

    # For Each Single AA-MM Excel Sheet
    for AA in AAMM_tabs:
        AAMM_tot[AA] = AA_total_count(SRTR, high_bins, low_bins, AA)
    
    # Calculate frequencies
    AAMM_freq = {AA: count / overall_SRTR_count for AA, count in AAMM_tot.items()}
    AAMM_freq_sorted = dict(sorted(AAMM_freq.items(), key=lambda item: item[1], reverse=True))

    # Generate datatables for AAMM frequency
    df_freq = pd.DataFrame(list(AAMM_freq_sorted.items()), columns=['AAMM', 'Frequency'])
    df_freq['Count'] = df_freq['AAMM'].map(AAMM_tot)

    # Add columns to indicate 'AA of Interest' and 'FIBERS Bin'
    df_freq['AA of Interest'] = df_freq['AAMM'].apply(lambda x: 1 if x in aa_of_int else 0)
    df_freq[f'FIBERS {bin_name}'] = df_freq['AAMM'].apply(lambda x: 1 if x in bin_of_int else 0)
    
    df_freq = df_freq[['AAMM', 'Count', 'Frequency', 'AA of Interest', f'FIBERS {bin_name}']]

    # Add a new worksheet for this population
    sheet = workbook.create_sheet(title=popcode)

    # Write the header
    sheet.append(['AAMM', 'Count', 'Frequency', 'AA of Interest', f'FIBERS {bin_name}'])

    # Write the DataFrame to the worksheet
    for r in df_freq.itertuples(index=False):
        sheet.append(r)

    # Plot AAMM frequencies
    fig, ax = plt.subplots(figsize=(14, 6))

    # Color code AAMM of interest vs. AAMM from FIBERS bin
    colors = {'AA of Interest': 'red', f'FIBERS {bin_name}': 'blue', 'Other': 'gray'}
    for i, row in df_freq.iterrows():
        if row['AA of Interest'] == 1 and row[f'FIBERS {bin_name}'] == 1:
            color = 'purple'  # or any color you prefer for AAs in both categories
            label = 'Both'
        elif row['AA of Interest'] == 1:
            color = colors['AA of Interest']
            label = 'AA of Interest'
        elif row[f'FIBERS {bin_name}'] == 1:
            color = colors[f'FIBERS {bin_name}']
            label = f'FIBERS {bin_name}'
        else:
            color = colors['Other']
            label = 'Other'
        
        ax.bar(row['AAMM'], row['Frequency'], color=color, label=label)

    # Remove duplicate labels
    handles, labels = plt.gca().get_legend_handles_labels()
    by_label = dict(zip(labels, handles))
    plt.legend(by_label.values(), by_label.keys())

    ax.set_title(f'AAMM Frequency Distribution Across Positions - {popcode}')
    ax.set_xlabel('AAMM')
    ax.set_ylabel('Frequency')

    # Adjust x-axis labels (font size)
    plt.xticks(rotation=45, ha='right', fontsize=6)

    # Increases the bottom margin of the plot to accommodate rotated labels
    #plt.subplots_adjust(bottom=0.2)

    #plt.legend()
    plt.tight_layout()
    plt.savefig(f'FIBERS_SRTR_AAMM_{bin_name}_{popcode}_freq_distr_plot_v3.png', dpi=300, bbox_inches='tight')
    plt.close()

    ######################

    with pd.ExcelWriter("FIBERS_SRTR_AAMM_" + bin_name + "_" + popcode + "_summary_table_v3.xlsx") as writer:
        OVERALL.to_excel(writer, sheet_name="OVERALL")
        table_df.to_excel(writer, sheet_name="AA-MM MEMBERSHIP")
        for AA in AAMM_tabs:
            AAMM_df[AA].to_excel(writer, sheet_name=AA + "_COUNT") # Saving to Excel removes duplicate sheet names

# Save datatables for each population group to a single excel spreadsheet
workbook.save(f'FIBERS_SRTR_AAMM_{bin_name}_freq_data_table_v3.xlsx')

# For Excel sheets
FIBERS_heading = NamedStyle(name="FIBERS_heading")
FIBERS_heading.font = Font(bold=True, color="00FF0000")

# Function for what each Excel sheet should look like
def excel_adjuster(worksheet, which_sheet):

    max_column = worksheet.max_column

    for column_index in range(1, max_column + 1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = 15

        for cell in worksheet[column_letter]:
            cell.alignment = Alignment(horizontal='left')

    if which_sheet == "TABLE":
        worksheet['A1'].value = "Top Bin For Each Imputation"
        worksheet['A1'].style = FIBERS_heading
        worksheet.column_dimensions['A'].width = 25
        for column_index in range(1, max_column + 1):
            column_letter = worksheet.cell(row=1, column=column_index).column_letter
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(horizontal='center')
    elif which_sheet == "OVERALL":
        worksheet['A1'].value = "MM Category"
    else:
        worksheet['A1'].value = "AA-MM within FIBERS MM Category"

    if which_sheet != "TABLE":
        worksheet.column_dimensions['A'].width = 38
        worksheet.move_range("A27:K27", rows=-24)
        worksheet.move_range("A50:K50", rows=-46)
        worksheet.move_range("A28:K49", rows=-1)
        worksheet['A1'].style = FIBERS_heading
        worksheet['A5'].style = FIBERS_heading
        worksheet['A27'].style = FIBERS_heading

    return worksheet


# Formatting the Excel sheets for readability
for pop in pops:

    popcode = popnames[pop]

    wb = load_workbook(filename="FIBERS_SRTR_AAMM_" + bin_name + "_" + popcode + "_summary_table_v3.xlsx")
    wb.add_named_style(FIBERS_heading)
    overall_ws = wb['OVERALL']
    table_ws = wb['AA-MM MEMBERSHIP']

    excel_adjuster(overall_ws, "OVERALL")
    excel_adjuster(table_ws, "TABLE")
    for AA in AAMM_tabs:
        excel_adjuster(wb[AA + "_COUNT"], "AA-MM")

    # Create a row for indicating bin of interest
    overall_ws.insert_rows(2)
    overall_ws['A2'].value = f"Analyzing FIBERS {bin_name}"
    overall_ws['A2'].style = FIBERS_heading

    # Create a row for a new heading for total population
    overall_ws.insert_rows(3)
    overall_ws['A3'].value = "Total Population"
    overall_ws['A3'].style = FIBERS_heading

    if pop == "OVERALL":
        overall_ws['A4'].value = "Total Population in SRTR Kidney"
        overall_ws.font = Font(bold=True)
    else:
        overall_ws['A4'].value = "Total " + popcode + " Population"
        overall_ws.font = Font(bold=True)

    for AA in AAMM_tabs:
        aa_ws = wb[AA + "_COUNT"]
        AA_SHEET_NAME = aa_ws.title.replace("_COUNT", "")

        # Create a row for indicating AA or bin of interest
        aa_ws.insert_rows(2)
        if AA_SHEET_NAME in aa_of_int and AA_SHEET_NAME in bin_of_int:
            aa_ws['A2'].value = f"AA of Interest and in FIBERS {bin_name}"
        elif AA_SHEET_NAME in aa_of_int:
            aa_ws['A2'].value = "AA of Interest"
        else:
            aa_ws["A2"].value = f"Position included in FIBERS {bin_name}"
        aa_ws["A2"].style = FIBERS_heading

        aa_ws.insert_rows(3)
        aa_ws['A3'].value = "Total " + AA + "_AAMM"
        aa_ws['A3'].style = FIBERS_heading
        if pop == "OVERALL":
            aa_ws['A4'].value = "Total Population with " + AA + "_AAMM"
            aa_ws.font = Font(bold=True)
        else:
            aa_ws['A4'].value = "Total " + popcode + " Population with " + AA + "_AAMM"
            aa_ws.font = Font(bold=True)

    wb.save(filename="FIBERS_SRTR_AAMM_" + bin_name + "_" + popcode + "_summary_table_v3.xlsx")

# Add a sheet for FIBERS run parameters
for pop in pops:

    popcode = popnames[pop]

    wb = load_workbook(filename="FIBERS_SRTR_AAMM_" + bin_name + "_" + popcode + "_summary_table_v3.xlsx")

    # Assuming you already have a workbook object named 'wb'
    sheet = wb.create_sheet("FIBERS_PARAMS")

    # Move the sheet to be the third sheet
    wb.move_sheet(sheet, offset=-len(wb.sheetnames)+3)

    # Write headers
    headers = ['FIBERS Bin', 'Dataset Filename', 'Random Seed', 'Bin Features', 
               'Threshold', 'Fitness', 'Pre-Fitness', 
               'Log-Rank Score', 'Log-Rank p-value', 'Bin Size', 'Group Ratio', 
               'Count At/Below Threshold', 'Count Above Threshold', 
               'Birth Iteration', 'Deletion Probability', 'Cluster', 
               'Residual', 'Residual p-value', 
               'Unadjusted HR', 'Unadjusted HR CI', 'Unadjusted HR p-value', 
               'Adjusted HR', 'Adjusted HR CI', 'Adjusted HR p-value', 
               'Runtime']
    sheet.append(headers)

    # Get the list of FIBERS Bins from the config file
    config_bins = set(config['FIBERS_BINS'].keys())

    # Filter bin_param to include only bins from the config file
    filtered_bin_param = [bin_data for bin_data in bin_param if bin_data['FIBERS Bin'] in config_bins]

    # Write data for each bin
    for bin_data in filtered_bin_param:
        sheet.append([bin_data[header] for header in headers])

    # Adjust column widths and wrap text
    for col in sheet.columns:
        column = col[0].column_letter  # Get the column name
        column_name = col[0].value  # Get the column header name

        if column_name == 'Bin Features':
            sheet.column_dimensions[column].width = 50
        else:
            max_length = 0
            for cell in col: # Don't skip the header cell
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) # * 1.2
            sheet.column_dimensions[column].width = min(adjusted_width, 50)  # Cap width at 50

        # Wrap text for all cells in the column
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Save the workbook
    wb.save(filename="FIBERS_SRTR_AAMM_" + bin_name + "_" + popcode + "_summary_table_v3.xlsx")
    
