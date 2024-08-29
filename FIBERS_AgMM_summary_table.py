import pandas as pd
import os
import sys
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook

# Script to generate summary stats for antigen MM and amino acid MM

# Overall Antigen Mismatch Counts Function
def antigen_count(SRTR, replicate, total_count):
    antigen_DR_NA_count = len(SRTR[SRTR['REC_DR_MM_EQUIV_CUR'] == 99])
    antigen_DQ_NA_count = len(SRTR[SRTR['REC_DQ_MM_EQUIV_CUR'] == 99])
    antigen_DR_NA_perc = str(antigen_DR_NA_count) + " (" + str(format(100 * antigen_DR_NA_count / total_count, '.2f')) + "%)"
    antigen_DQ_NA_perc = str(antigen_DQ_NA_count) + " (" + str(format(100 * antigen_DQ_NA_count / total_count, ".2f")) + "%)"

    antigen_ABDR_0MM_count = len(SRTR[(SRTR['REC_A_MM_EQUIV_CUR'] == 0) &
                                      (SRTR['REC_B_MM_EQUIV_CUR'] == 0) &
                                      (SRTR['REC_DR_MM_EQUIV_CUR'] == 0)])
    antigen_ABDR_0MM_perc = str(antigen_ABDR_0MM_count) + " (" + str(
        format(100 * antigen_ABDR_0MM_count / total_count, '.2f')) + "%)"

    antigen_DR_0MM_count = len(SRTR[SRTR['REC_DR_MM_EQUIV_CUR'] == 0])
    antigen_DQ_0MM_count = len(SRTR[SRTR['REC_DQ_MM_EQUIV_CUR'] == 0])
    antigen_DR_0MM_perc = str(antigen_DR_0MM_count) + " (" + str(format(100 * antigen_DR_0MM_count / total_count, '.2f')) + "%)"
    antigen_DQ_0MM_perc = str(antigen_DQ_0MM_count) + " (" + str(format(100 * antigen_DQ_0MM_count / total_count, '.2f')) + "%)"

    antigen_DR_1MM_count = len(SRTR[SRTR['REC_DR_MM_EQUIV_CUR'] == 1])
    antigen_DQ_1MM_count = len(SRTR[SRTR['REC_DQ_MM_EQUIV_CUR'] == 1])
    antigen_DR_1MM_perc = str(antigen_DR_1MM_count) + " (" + str(format(100 * antigen_DR_1MM_count / total_count, '.2f')) + "%)"
    antigen_DQ_1MM_perc = str(antigen_DQ_1MM_count) + " (" + str(format(100 * antigen_DQ_1MM_count / total_count, '.2f')) + "%)"

    antigen_DR_2MM_count = len(SRTR[SRTR['REC_DR_MM_EQUIV_CUR'] == 2])
    antigen_DQ_2MM_count = len(SRTR[SRTR['REC_DQ_MM_EQUIV_CUR'] == 2])
    antigen_DR_2MM_perc = str(antigen_DR_2MM_count) + " (" + str(format(100 * antigen_DR_2MM_count / total_count, '.2f')) + "%)"
    antigen_DQ_2MM_perc = str(antigen_DQ_2MM_count) + " (" + str(format(100 * antigen_DQ_2MM_count / total_count, '.2f')) + "%)"

    antigen_DR_0MM_DQ_0MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 0) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 0)])
    antigen_DR_0MM_DQ_0MM_perc = str(antigen_DR_0MM_DQ_0MM_count) + " (" + str(
        format(100 * antigen_DR_0MM_DQ_0MM_count / total_count, '.2f')) + "%)"

    antigen_DR_0MM_DQ_1MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 0) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 1)])
    antigen_DR_0MM_DQ_1MM_perc = str(antigen_DR_0MM_DQ_1MM_count) + " (" + str(
        format(100 * antigen_DR_0MM_DQ_1MM_count / total_count, '.2f')) + "%)"

    antigen_DR_1MM_DQ_0MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 1) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 0)])
    antigen_DR_1MM_DQ_0MM_perc = str(antigen_DR_1MM_DQ_0MM_count) + " (" + str(
        format(100 * antigen_DR_1MM_DQ_0MM_count / total_count, '.2f')) + "%)"

    antigen_DR_1MM_DQ_1MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 1) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 1)])
    antigen_DR_1MM_DQ_1MM_perc = str(antigen_DR_1MM_DQ_1MM_count) + " (" + str(
        format(100 * antigen_DR_1MM_DQ_1MM_count / total_count, '.2f')) + "%)"

    antigen_DR_0MM_DQ_2MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 0) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 2)])
    antigen_DR_0MM_DQ_2MM_perc = str(antigen_DR_0MM_DQ_2MM_count) + " (" + str(
        format(100 * antigen_DR_0MM_DQ_2MM_count / total_count, '.2f')) + "%)"

    antigen_DR_2MM_DQ_0MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 2) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 0)])
    antigen_DR_2MM_DQ_0MM_perc = str(antigen_DR_2MM_DQ_0MM_count) + " (" + str(
        format(100 * antigen_DR_2MM_DQ_0MM_count / total_count, '.2f')) + "%)"

    antigen_DR_1MM_DQ_2MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 1) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 2)])
    antigen_DR_1MM_DQ_2MM_perc = str(antigen_DR_1MM_DQ_2MM_count) + " (" + str(
        format(100 * antigen_DR_1MM_DQ_2MM_count / total_count, '.2f')) + "%)"

    antigen_DR_2MM_DQ_1MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 2) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 1)])
    antigen_DR_2MM_DQ_1MM_perc = str(antigen_DR_2MM_DQ_1MM_count) + " (" + str(
        format(100 * antigen_DR_2MM_DQ_1MM_count / total_count, '.2f')) + "%)"

    antigen_DR_2MM_DQ_2MM_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 2) &
                                           (SRTR['REC_DQ_MM_EQUIV_CUR'] == 2)])
    antigen_DR_2MM_DQ_2MM_perc = str(antigen_DR_2MM_DQ_2MM_count) + " (" + str(
        format(100 * antigen_DR_2MM_DQ_2MM_count / total_count, '.2f')) + "%)"

    antigen_DR_0MM_DQ_NA_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 0) &
                                          (SRTR['REC_DQ_MM_EQUIV_CUR'] == 99)])
    antigen_DR_0MM_DQ_NA_perc = str(antigen_DR_0MM_DQ_NA_count) + " (" + str(
        format(100 * antigen_DR_0MM_DQ_NA_count / total_count, '.2f')) + "%)"

    antigen_DR_1MM_DQ_NA_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 1) &
                                          (SRTR['REC_DQ_MM_EQUIV_CUR'] == 99)])
    antigen_DR_1MM_DQ_NA_perc = str(antigen_DR_1MM_DQ_NA_count) + " (" + str(
        format(100 * antigen_DR_1MM_DQ_NA_count / total_count, '.2f')) + "%)"

    antigen_DR_2MM_DQ_NA_count = len(SRTR[(SRTR['REC_DR_MM_EQUIV_CUR'] == 2) &
                                          (SRTR['REC_DQ_MM_EQUIV_CUR'] == 99)])
    antigen_DR_2MM_DQ_NA_perc = str(antigen_DR_2MM_DQ_NA_count) + " (" + str(
        format(100 * antigen_DR_2MM_DQ_NA_count / total_count, '.2f')) + "%)"

    antigen_list = [antigen_DR_NA_perc, antigen_DQ_NA_perc, antigen_ABDR_0MM_perc,
                    antigen_DR_0MM_perc, antigen_DQ_0MM_perc, antigen_DR_1MM_perc,
                    antigen_DQ_1MM_perc, antigen_DR_2MM_perc, antigen_DQ_2MM_perc,
                    antigen_DR_0MM_DQ_0MM_perc, antigen_DR_0MM_DQ_1MM_perc,
                    antigen_DR_1MM_DQ_0MM_perc, antigen_DR_1MM_DQ_1MM_perc,
                    antigen_DR_0MM_DQ_2MM_perc, antigen_DR_2MM_DQ_0MM_perc,
                    antigen_DR_1MM_DQ_2MM_perc, antigen_DR_2MM_DQ_1MM_perc,
                    antigen_DR_2MM_DQ_2MM_perc, antigen_DR_0MM_DQ_NA_perc,
                    antigen_DR_1MM_DQ_NA_perc, antigen_DR_2MM_DQ_NA_perc]

    antigen_index = ["Missing-DR AgMM Count", "Missing-DQ AgMM Count", "0-ABDR AgMM Count", "0-DR AgMM Count",
                     "0-DQ AgMM Count", "1-DR AgMM Count", "1-DQ AgMM Count", "2-DR AgMM Count", "2-DQ AgMM Count",
                     "0-DR and 0-DQ AgMM Count", "0-DR and 1-DQ AgMM Count", "1-DR and 0-DQ AgMM Count",
                     "1-DR and 1-DQ AgMM Count", "0-DR and 2-DQ AgMM Count", "2-DR and 0-DQ AgMM Count",
                     "1-DR and 2-DQ AgMM Count", "2-DR and 1-DQ AgMM Count", "2-DR and 2-DQ AgMM Count",
                     "0-DR and Missing-DQ AgMM Count", "1-DR and Missing-DQ AgMM Count",
                     "2-DR and Missing-DQ AgMM Count"]
    number = replicate

    antigen_dataframe = pd.DataFrame(data=antigen_list, columns=[number], index=antigen_index)
    return antigen_dataframe


# Initialize DataFrames which will be each Excel sheet
headings = ["Replicate_1", "Replicate_2", "Replicate_3", "Replicate_4", "Replicate_5", "Replicate_6", "Replicate_7",
            "Replicate_8", "Replicate_9", "Replicate_10"]
OVERALL = pd.DataFrame(columns=headings)
FIBERS_HIGH = pd.DataFrame(columns=headings)
FIBERS_LOW = pd.DataFrame(columns=headings)

FIBERS_AAMM_bin = "BIN_9"           # Only BIN_9 for now
impute_realization = range(1, 11)   # 1 thru 10 for each realization/replicate

pops = ["OVERALL", "AFA", "CAU", "HIS", "HPI", "NAM", "ASI", "Multi-Racial"]

for pop in pops:
    print("Working on population: ", pop)
    for replicate in impute_realization:
        print("Working on replicate: ", replicate)
        # SRTR imputation replicate
        SRTR_imputation_replicate_filename = "SRTR_AA_MM_9loc_matrix_" + str(replicate) + ".txt.gz"
        SRTR = pd.read_csv(SRTR_imputation_replicate_filename, sep='\t', compression='gzip')

        # subset antigen MM and amino acid MM columns from data frame

        SRTR = SRTR[
            ["PX_ID", "CAN_RACE", "REC_A_MM_EQUIV_CUR", "REC_B_MM_EQUIV_CUR", "REC_DR_MM_EQUIV_CUR",
             "REC_DQ_MM_EQUIV_CUR",
             "MM_DRB1_13", "MM_DRB1_26", "MM_DQB1_30", "MM_DQB1_55",
             "MM_DRB1_11", "MM_DRB1_37", "MM_DQB1_53", "MM_DRB1_77", "MM_DQB1_89"
             ]]

        # Convert AgMM to integers and handle NAs due to missing DQ typing
        # 99 - Not tested per SRTR data dictionary
        SRTR['REC_DR_MM_EQUIV_CUR'] = SRTR['REC_DR_MM_EQUIV_CUR'].fillna(99).astype(int)
        SRTR['REC_DQ_MM_EQUIV_CUR'] = SRTR['REC_DQ_MM_EQUIV_CUR'].fillna(99).astype(int)
        SRTR = SRTR.astype({'REC_DR_MM_EQUIV_CUR': 'int', 'REC_DQ_MM_EQUIV_CUR': 'int'})

        if pop == "OVERALL":
            SRTR_pop = SRTR
        else:
            SRTR_pop = SRTR[(SRTR['CAN_RACE'] == pop)]

        # For BIN_9 ONLY. FIBERS - AAMM Positions in Bin_9: DRB1_13, DRB1_26, DQB1_30, DQB1_55
        SRTR_Low_Risk = SRTR_pop[(SRTR_pop['MM_DRB1_13'] == 0) &
                             (SRTR_pop['MM_DRB1_26'] == 0) &
                             (SRTR_pop['MM_DQB1_30'] == 0) &
                             (SRTR_pop['MM_DQB1_55'] == 0)]  # For FIBERS_LOW df

        SRTR_High_Risk = SRTR_pop[(SRTR_pop['MM_DRB1_13'] >= 1) |
                              (SRTR_pop['MM_DRB1_26'] >= 1) |
                              (SRTR_pop['MM_DQB1_30'] >= 1) |
                              (SRTR_pop['MM_DQB1_55'] >= 1)]  # For FIBERS_HIGH df

        SRTR_Overall_Risk = SRTR_pop  # For OVERALL df

        # overall count for each df
        kidney_tx_pair_count = len(SRTR_Overall_Risk)
        FIBERS_Low_Risk_count = len(SRTR_Low_Risk)
        FIBERS_High_Risk_count = len(SRTR_High_Risk)

        # Call the function for each dataframe
        overall_antigen = antigen_count(SRTR_Overall_Risk, replicate, kidney_tx_pair_count)
        OVERALL["Replicate_" + str(replicate)] = overall_antigen
        low_antigen = antigen_count(SRTR_Low_Risk, replicate, FIBERS_Low_Risk_count)
        FIBERS_LOW["Replicate_" + str(replicate)] = low_antigen
        high_antigen = antigen_count(SRTR_High_Risk, replicate, FIBERS_High_Risk_count)
        FIBERS_HIGH["Replicate_" + str(replicate)] = high_antigen

    with pd.ExcelWriter("FIBERS_AgMM_" + pop + "_summary_table.xlsx") as writer:
        OVERALL.to_excel(writer, sheet_name="OVERALL")
        FIBERS_HIGH.to_excel(writer, sheet_name="FIBERS_HIGH")
        FIBERS_LOW.to_excel(writer, sheet_name="FIBERS_LOW")
        
        
# Formatting the Excel sheets for readability
for pop in pops:
    wb = load_workbook(filename="FIBERS_AgMM_" + pop + "_summary_table.xlsx")
    overall_ws = wb['OVERALL']
    high_ws = wb['FIBERS_HIGH']
    low_ws = wb['FIBERS_LOW']

    overall_ws['A1'].value = "Mismatch Category"
    overall_ws['A1'].font = Font(bold=True)
    high_ws['A1'].value = "Mismatch Category"
    high_ws['A1'].font = Font(bold=True)
    low_ws['A1'].value = "Mismatch Category"
    low_ws['A1'].font = Font(bold=True)

    max_column = overall_ws.max_column

    # Increase the width and left-alignment each column
    for column_index in range(1, max_column + 1):
        column_letter = overall_ws.cell(row=1, column=column_index).column_letter
        overall_ws.column_dimensions[column_letter].width = 14
        high_ws.column_dimensions[column_letter].width = 14
        low_ws.column_dimensions[column_letter].width = 14

        for cell in overall_ws[column_letter]:
            cell.alignment = Alignment(horizontal='left')
        for cell in high_ws[column_letter]:
            cell.alignment = Alignment(horizontal='left')
        for cell in low_ws[column_letter]:
            cell.alignment = Alignment(horizontal='left')

    overall_ws.column_dimensions['A'].width = 30
    high_ws.column_dimensions['A'].width = 30
    low_ws.column_dimensions['A'].width = 30

    wb.save(filename="FIBERS_AgMM_" + pop + "_summary_table.xlsx")



  


