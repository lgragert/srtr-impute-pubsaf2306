import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import hlagenie
genie = hlagenie.init("3510", imputed=True)

# Go from serologic group to allele to amino acid pairings


# Get amino acid position with hlagenie
def aminoacid(ag_allele, which_allele):

    # Get length of protein sequence
    if which_allele == 'A':
        AA_list = list(range(1,342))
    elif which_allele == 'C':
        AA_list = list(range(1,343))
    elif which_allele == 'B':
        AA_list = list(range(1,339))
    elif which_allele == 'DRB1' or which_allele == 'DRB345':
        AA_list = list(range(1,238))
    elif which_allele == 'DQA1':
        AA_list = list(range(1, 233))
    elif which_allele == 'DQB1' or which_allele == 'DPA1' or which_allele == 'DPB1':
        AA_list = list(range(1, 230))

    for AA in AA_list:
        locus_AA = []
        for row in range(len(ag_allele)):
            allele = ag_allele['HLA_Allele'][row]
            # Ask if these are ok substitutions
            if allele == 'C*06:100':
                allele = 'C*06:10'
            if allele == 'B*15:245Q':
                allele = 'B*15:220'
            if allele == 'DQA1*03:07':
                allele = 'DQA1*03:06'
            if allele == 'DPB1*1038:01':
                allele = 'DPB1*01:01'
            locus_AA.append(genie.getAA(allele, AA))
        ag_allele[which_allele + "_" + str(AA)] = locus_AA
    return ag_allele


# Transpose dataset to get the data to be what you want
def get_allele_format(optn_ag, ag_alleles):

    # Only keep the first ten DR alleles for now
    ag_alleles = ag_alleles['IMGT/HLA alleles'].str.split(pat='/', expand=True)
    ag_alleles = ag_alleles.iloc[:, :10]
    ag_alleles = pd.concat([optn_ag, ag_alleles], axis=1).reset_index(drop=True)

    # Want the rows to be indexed by allele
    transpose_ag_allele = ag_alleles.transpose()
    transpose_ag_allele.columns = transpose_ag_allele.iloc[0]
    transpose_ag_allele = transpose_ag_allele.iloc[1:]
    melt_ag_allele = pd.melt(transpose_ag_allele)
    melt_ag_allele = melt_ag_allele.dropna(ignore_index=True)
    melt_ag_allele = melt_ag_allele.rename(columns={'value': 'HLA_Allele'})
    return melt_ag_allele


# Two blank lines between each serological group
def blank_lines_insert(allele_aa, optn_ag_list):
    index_list = []
    for serology in optn_ag_list:
        index_list = allele_aa[(allele_aa["OPTN_Antigen"] == serology)].index.values
        # last_index_str = ''.join(map(str, index_list[-1:]))
        # last_index = int(last_index_str)
        last_index = index_list[-1]
        blank_line = pd.DataFrame(index=[0, 1], columns=allele_aa.columns)
        allele_aa = pd.concat([allele_aa.loc[:last_index], blank_line, allele_aa.loc[last_index + 1:]]).reset_index(drop=True)
    allele_aa = allele_aa[:-2]
    return allele_aa


# First start with list of allele serology with antigen type
ag_to_alleles = pd.read_csv('OPTN_antigens_to_alleles_CPRA.txt', sep='\t')

loci = ['A', 'C', 'B', 'DRB', 'DQA1', 'DQB1', 'DPA1', 'DPB1']

# Do all loci with their AAs, each sheet is one locus
for locus in loci:

    # Want to separate out DRB into DRB1 and DRB345
    if locus == 'DRB':
        DRB1_ag_alleles = ag_to_alleles[(ag_to_alleles['IMGT/HLA alleles'].str.startswith('DRB1'))]
        DRB345_ag_alleles = ag_to_alleles[(ag_to_alleles['IMGT/HLA alleles'].str.startswith('DRB3')) |
                                          (ag_to_alleles['IMGT/HLA alleles'].str.startswith('DRB4')) |
                                          (ag_to_alleles['IMGT/HLA alleles'].str.startswith('DRB5'))]

        DRB1_optn_ag = DRB1_ag_alleles['OPTN_Antigen']
        DRB1_optn_ag_ls = DRB1_optn_ag.values.tolist()
        alleles = get_allele_format(DRB1_optn_ag, DRB1_ag_alleles)
        DRB1_aa = aminoacid(alleles, 'DRB1')

        DRB345_optn_ag = DRB345_ag_alleles['OPTN_Antigen']
        DRB345_optn_ag_ls = DRB345_optn_ag.values.tolist()
        alleles = get_allele_format(DRB345_optn_ag, DRB345_ag_alleles)
        DRB345_aa = aminoacid(alleles, 'DRB345')

        # Putting blank lines inbetween each serologic group
        DRB1_aa = blank_lines_insert(DRB1_aa, DRB1_optn_ag_ls)
        DRB345_aa = blank_lines_insert(DRB345_aa, DRB345_optn_ag_ls)
        with pd.ExcelWriter("AA_polymorphism_with_allele.xlsx", mode="a", engine="openpyxl",
                            if_sheet_exists="replace") as writer:
            DRB1_aa.to_excel(writer, index=False, sheet_name='DRB1 Alleles & AA')
            DRB345_aa.to_excel(writer, index=False, sheet_name='DRB345 Alleles & AA')
    else:
        ag_alleles = ag_to_alleles[(ag_to_alleles['IMGT/HLA alleles'].str.startswith(locus))]
        locus_optn_ag = ag_alleles['OPTN_Antigen']
        locus_optn_ag_ls = locus_optn_ag.values.tolist()

        alleles = get_allele_format(locus_optn_ag, ag_alleles)

        # Use HLAGenie to get the AA positions
        allele_aa = aminoacid(alleles, locus)

        # Putting blank lines inbetween each serologic group
        allele_aa = blank_lines_insert(allele_aa, locus_optn_ag_ls)

        with pd.ExcelWriter("AA_polymorphism_with_allele.xlsx", mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
            allele_aa.to_excel(writer, index=False, sheet_name=locus + ' Alleles & AA')


# Use to get amino acid color for each worksheet
def amino_acid_color(worksheet):
    max_column = worksheet.max_column
    for col in range(2, max_column + 1):
        column_letter = worksheet.cell(row=2, column=col).column_letter
        for cell in worksheet[column_letter]:
            if cell.value in AA_colors.keys():
                color_AA = PatternFill(patternType='solid', fgColor=AA_colors.get(cell.value))
                cell.fill = color_AA
            else:
                continue

    worksheet.column_dimensions['B'].width = 13
    return worksheet


wb = load_workbook("AA_polymorphism_with_allele.xlsx")
loci = ['A', 'C', 'B', 'DRB1', 'DRB345', 'DQA1', 'DQB1', 'DPA1', 'DPB1']
for locus in loci:
    ws_locus = wb[locus + " Alleles & AA"]

    # Spring color scheme
    AA_colors = {"A": "ffababab", "C": "ff0893fe", "D": "fffc5977", "E": "ffff9c83", "F": "ff02f8b7", "G": "fff66c17",
                 "H": "ffd2b907", "I": "ff47beff", "K": "ffffc56d", "L": "ff358b8f", "M": "ff4b9e93", "N": "ffeb684b",
                 "P": "fffd75cf", "Q": "ffdeb16f", "R": "ffb38b2b", "S": "ffbb8187", "T": "ffd3add1", "V": "ff519db9",
                 "W": "ff00d619", "Y": "ff38a549"}

    ws_locus = amino_acid_color(ws_locus)

wb.save(filename="AA_polymorphism_with_allele.xlsx")
